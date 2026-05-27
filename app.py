import sqlite3
import face_recognition
import numpy as np
import base64
import io
import cv2
import os
from PIL import Image
from flask import Flask, render_template, request, redirect, url_for, Response, session, send_file
from datetime import datetime
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.secret_key = "gns_vision_secure_key_2026"

UPLOAD_FOLDER = os.path.join('static', 'timetables')
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
os.makedirs(UPLOAD_FOLDER, exist_ok=True)


def init_db():
    with sqlite3.connect('attendance_system.db') as conn:
        cursor = conn.cursor()
        cursor.execute('''CREATE TABLE IF NOT EXISTS students 
                          (id INTEGER PRIMARY KEY AUTOINCREMENT, 
                           roll_id TEXT UNIQUE, name TEXT, encoding TEXT, 
                           class_name TEXT, is_authorized INTEGER DEFAULT 0)''')

        cursor.execute('''CREATE TABLE IF NOT EXISTS attendance 
                          (id INTEGER PRIMARY KEY AUTOINCREMENT, 
                           roll_id TEXT, subject TEXT, date TEXT, time TEXT, time_slot TEXT, status TEXT,
                           UNIQUE(roll_id, date, subject, time_slot))''')

        cursor.execute('''CREATE TABLE IF NOT EXISTS timetable_images 
                          (id INTEGER PRIMARY KEY AUTOINCREMENT, class_name TEXT UNIQUE, image_path TEXT)''')

        cursor.execute('''CREATE TABLE IF NOT EXISTS teachers 
                          (id INTEGER PRIMARY KEY AUTOINCREMENT, username TEXT UNIQUE, password TEXT)''')
        try:
            cursor.execute("INSERT INTO teachers (username, password) VALUES (?, ?)", ('admin', 'admin123'))
        except sqlite3.IntegrityError:
            pass
        conn.commit()


init_db()


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        with sqlite3.connect('attendance_system.db') as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM teachers WHERE username=? AND password=?", (username, password))
            if cursor.fetchone():
                session['logged_in'] = True
                return redirect(url_for('admin'))
        return "<script>alert('Invalid Credentials!'); window.location='/login';</script>"
    return render_template('teacher_login.html')


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('index'))


@app.route('/admin')
def admin():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    with sqlite3.connect('attendance_system.db') as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM students")
        students = cursor.fetchall()
        cursor.execute("SELECT * FROM timetable_images")
        timetables = cursor.fetchall()
    return render_template('admin.html', students=students, timetables=timetables)


@app.route('/upload_timetable', methods=['POST'])
def upload_timetable():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    class_name = request.form['class_name']
    file = request.files.get('timetable_image')
    if file and file.filename != '':
        filename = secure_filename(f"{class_name}_{file.filename}")
        file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
        db_path = f"timetables/{filename}"
        with sqlite3.connect('attendance_system.db') as conn:
            cursor = conn.cursor()
            cursor.execute('''INSERT INTO timetable_images (class_name, image_path) VALUES (?, ?)
                              ON CONFLICT(class_name) DO UPDATE SET image_path=excluded.image_path''',
                           (class_name, db_path))
            conn.commit()
        return "<script>alert('Timetable Uploaded!'); window.location='/admin';</script>"
    return "<script>alert('Upload Failed!'); window.location='/admin';</script>"


@app.route('/delete_timetable_image/<int:id>')
def delete_timetable_image(id):
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    with sqlite3.connect('attendance_system.db') as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT image_path FROM timetable_images WHERE id = ?", (id,))
        row = cursor.fetchone()
        if row:
            full_path = os.path.join('static', row[0])
            if os.path.exists(full_path):
                os.remove(full_path)
            cursor.execute("DELETE FROM timetable_images WHERE id = ?", (id,))
            conn.commit()
    return "<script>alert('Timetable Removed!'); window.location='/admin';</script>"


@app.route('/approve/<int:id>')
def approve(id):
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    with sqlite3.connect('attendance_system.db') as conn:
        conn.cursor().execute("UPDATE students SET is_authorized = 1 WHERE id = ?", (id,))
        conn.commit()
    return redirect(url_for('admin'))


@app.route('/delete_student/<int:id>')
def delete_student(id):
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    with sqlite3.connect('attendance_system.db') as conn:
        conn.cursor().execute("DELETE FROM students WHERE id = ?", (id,))
        conn.commit()
    return redirect(url_for('admin'))


@app.route('/attendance', methods=['GET', 'POST'])
def attendance():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    if request.method == 'POST':
        class_name = request.form['class_name']
        subject = request.form['subject']
        time_slot = request.form['time_slot'].strip()
        current_date = datetime.now().strftime("%Y-%m-%d")

        with sqlite3.connect('attendance_system.db') as conn:
            cursor = conn.cursor()
            cursor.execute('''SELECT COUNT(*) FROM attendance 
                              WHERE date=? AND time_slot=? AND roll_id IN 
                              (SELECT roll_id FROM students WHERE class_name=?)''',
                           (current_date, time_slot, class_name))
            if cursor.fetchone()[0] > 0:
                return f"<script>alert('Conflict Lock: Attendance already recorded for this class in time slot ({time_slot}) today!'); window.location='/attendance';</script>"
        return render_template('attendance.html', class_name=class_name, subject=subject, time_slot=time_slot)
    return render_template('select_subject.html')


@app.route('/shutdown_session/<class_name>/<subject>/<time_slot>')
def shutdown_session(class_name, subject, time_slot):
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    current_date = datetime.now().strftime("%Y-%m-%d")
    current_time = datetime.now().strftime("%H:%M:%S")
    with sqlite3.connect('attendance_system.db') as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT roll_id FROM students WHERE class_name = ? AND is_authorized = 1", (class_name,))
        students = cursor.fetchall()
        for s in students:
            try:
                cursor.execute('''INSERT INTO attendance (roll_id, subject, date, time, time_slot, status) 
                                  VALUES (?, ?, ?, ?, ?, 'Absent')''',
                               (s[0], subject, current_date, current_time, time_slot))
            except sqlite3.IntegrityError:
                pass
        conn.commit()
    return redirect(
        url_for('manage_attendance', class_name=class_name, subject=subject, time_slot=time_slot, date=current_date))


@app.route('/manage_attendance', methods=['GET', 'POST'])
def manage_attendance():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    class_name = request.form.get('class_name', request.args.get('class_name', 'BCA_3rd_Year'))
    subject = request.form.get('subject', request.args.get('subject', 'Python'))
    time_slot = request.form.get('time_slot', request.args.get('time_slot', '09:15 - 10:15')).strip()
    date = request.form.get('date', request.args.get('date', datetime.now().strftime("%Y-%m-%d")))

    today = datetime.now().strftime("%Y-%m-%d")
    if date > today:
        return "<script>alert('Future dates cannot be processed!'); window.location='/manage_attendance';</script>"

    is_editable = False
    if date == today and datetime.now().hour < 19:
        is_editable = True

    with sqlite3.connect('attendance_system.db') as conn:
        cursor = conn.cursor()
        cursor.execute('''SELECT s.roll_id, s.name, s.class_name, COALESCE(a.status, 'Absent'), COALESCE(a.time, '--:--:--'), ?
                          FROM students s LEFT JOIN attendance a ON s.roll_id = a.roll_id 
                          AND a.subject = ? AND a.date = ? AND a.time_slot = ?
                          WHERE s.class_name = ? AND s.is_authorized = 1''',
                       (time_slot, subject, date, time_slot, class_name))
        records = cursor.fetchall()
    return render_template('manage_attendance.html', records=records, class_name=class_name, subject=subject,
                           time_slot=time_slot, date=date, is_editable=is_editable)


@app.route('/toggle_status/<roll_id>/<class_name>/<subject>/<date>/<time_slot>')
def toggle_status(roll_id, class_name, subject, date, time_slot):
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    if date != datetime.now().strftime("%Y-%m-%d") or datetime.now().hour >= 19:
        return "<script>alert('Security Lock: Modifications denied after 7:00 PM.'); window.location='/manage_attendance';</script>"

    with sqlite3.connect('attendance_system.db') as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT status FROM attendance WHERE roll_id=? AND subject=? AND date=? AND time_slot=?",
                       (roll_id, subject, date, time_slot))
        row = cursor.fetchone()
        if row:
            new_status = "Absent" if row[0] == "Present" else "Present"
            cursor.execute("UPDATE attendance SET status=? WHERE roll_id=? AND subject=? AND date=? AND time_slot=?",
                           (new_status, roll_id, subject, date, time_slot))
        else:
            cursor.execute(
                "INSERT INTO attendance (roll_id, subject, date, time, time_slot, status) VALUES (?, ?, ?, ?, ?, 'Present')",
                (roll_id, subject, date, datetime.now().strftime("%H:%M:%S"), time_slot))
        conn.commit()
    return redirect(
        url_for('manage_attendance', class_name=class_name, subject=subject, date=date, time_slot=time_slot))


@app.route('/download_report/<class_name>')
def download_report(class_name):
    if not session.get('logged_in'): return redirect(url_for('login'))
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    with sqlite3.connect('attendance_system.db') as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT roll_id, name FROM students WHERE class_name = ? AND is_authorized = 1", (class_name,))
        students = cursor.fetchall()

    if not students:
        return "<script>alert('No authorized students found in this class!'); window.location='/admin';</script>"

    report_data = []
    with sqlite3.connect('attendance_system.db') as conn:
        cursor = conn.cursor()
        for roll_id, name in students:
            cursor.execute("SELECT COUNT(*) FROM attendance WHERE roll_id = ?", (roll_id,))
            total_lectures = cursor.fetchone()[0]
            cursor.execute("SELECT COUNT(*) FROM attendance WHERE roll_id = ? AND status = 'Present'", (roll_id,))
            attended_lectures = cursor.fetchone()[0]

            percentage = round((attended_lectures / total_lectures * 100), 2) if total_lectures > 0 else 0.0
            eligibility = "Eligible" if percentage >= 75.0 else "Debarred (<75%)"

            report_data.append({
                "Roll ID": roll_id, "Student Name": name, "Total Lectures": total_lectures,
                "Lectures Attended": attended_lectures, "Attendance Percentage (%)": percentage,
                "Eligibility Status": eligibility
            })

    df = pd.DataFrame(report_data)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="Attendance Summary")
    output.seek(0)

    wb = openpyxl.load_workbook(output)
    ws = wb.active

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    zebra_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    debarred_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    eligible_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    thin_border = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
                         top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))

    for cell in ws[1]:
        cell.fill, cell.font, cell.alignment = header_fill, header_font, Alignment(horizontal="center",
                                                                                   vertical="center")

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6), start=2):
        percentage_val = row[4].value
        is_even = (row_idx % 2 == 0)
        for cell in row:
            cell.font, cell.border, cell.alignment = Font(name="Calibri", size=11), thin_border, Alignment(
                horizontal="left", vertical="center")
            if is_even: cell.fill = zebra_fill

        row[0].alignment = row[2].alignment = row[3].alignment = row[4].alignment = row[5].alignment = Alignment(
            horizontal="center")
        if percentage_val < 75.0:
            row[5].fill, row[5].font = debarred_fill, Font(name="Calibri", size=11, bold=True, color="C00000")
        else:
            row[5].fill, row[5].font = eligible_fill, Font(name="Calibri", size=11, bold=True, color="375623")

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = max(max_len + 3, 12)

    final_output = io.BytesIO()
    wb.save(final_output)
    final_output.seek(0)

    return send_file(final_output, as_attachment=True, download_name=f"Attendance_Cumulative_{class_name}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route('/view_timetable', methods=['GET', 'POST'])
def view_timetable():
    selected_class = request.form.get('class_name') if request.method == 'POST' else None
    image_file = None
    if selected_class:
        with sqlite3.connect('attendance_system.db') as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT image_path FROM timetable_images WHERE class_name = ?", (selected_class,))
            row = cursor.fetchone()
            if row: image_file = row[0]
    return render_template('view_timetable.html', selected_class=selected_class, image_file=image_file)


@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        name, roll_id, class_name = request.form['name'], request.form['roll_id'], request.form['class_name']
        image_data = request.form['image_data']
        try:
            encoded = image_data.split(",", 1)[1]
            img = Image.open(io.BytesIO(base64.b64decode(encoded)))
            img_np = np.array(img.convert('RGB'))
            face_locs = face_recognition.face_locations(img_np)
            encs = face_recognition.face_encodings(img_np, face_locs)

            if len(encs) > 0:
                new_encoding = encs[0]
                with sqlite3.connect('attendance_system.db') as conn:
                    cursor = conn.cursor()
                    cursor.execute("SELECT roll_id, encoding FROM students")
                    db_students = cursor.fetchall()

                if db_students:
                    known_encodings = [np.fromstring(row[1], sep=",") for row in db_students]
                    known_rolls = [row[0] for row in db_students]
                    duplicates = face_recognition.compare_faces(known_encodings, new_encoding, 0.45)
                    if True in duplicates:
                        existing_roll = known_rolls[duplicates.index(True)]
                        return f"<script>alert('Blunder Blocked: Registered under Roll ID: {existing_roll}!'); window.location='/register';</script>"

                enc_str = ",".join(map(str, new_encoding))
                with sqlite3.connect('attendance_system.db') as conn:
                    conn.cursor().execute(
                        "INSERT INTO students (roll_id, name, encoding, class_name) VALUES (?, ?, ?, ?)",
                        (roll_id, name, enc_str, class_name))
                return "<script>alert('Registration Successful!'); window.location='/';</script>"
            return "<script>alert('No face detected! Try again.'); window.location='/register';</script>"
        except Exception as e:
            return f"<script>alert('Error: {str(e)}'); window.location='/register';</script>"
    return render_template('register.html')


@app.route('/dashboard', methods=['GET', 'POST'])
def dashboard():
    if request.method == 'POST':
        roll_id = request.form.get('roll_id').strip()
        with sqlite3.connect('attendance_system.db') as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT name, class_name FROM students WHERE roll_id = ?", (roll_id,))
            student = cursor.fetchone()
            if student:
                cursor.execute("SELECT subject, date, time, status, time_slot FROM attendance WHERE roll_id = ?",
                               (roll_id,))
                history = cursor.fetchall()
                present = len([h for h in history if h[3] == 'Present'])
                total = len(history)
                perc = (present / total * 100) if total > 0 else 0
                return render_template('dashboard.html', name=student[0], class_name=student[1], roll=roll_id,
                                       present=present, total=total, perc=round(perc, 2), history=history)
        return "<script>alert('Roll ID records missing!'); window.location='/dashboard';</script>"
    return render_template('login.html')


# === BINDING FIX: Endpoint function name mapped explicitly as 'video_feed' ===
@app.route('/video_feed/<class_name>/<subject>/<time_slot>')
def video_feed(class_name, subject, time_slot):
    return Response(video_stream(class_name, subject, time_slot), mimetype='multipart/x-mixed-replace; boundary=frame')


def video_stream(class_name, subject, time_slot):
    camera = cv2.VideoCapture(0, cv2.CAP_DSHOW)
    if not camera.isOpened(): camera = cv2.VideoCapture(0)
    try:
        while True:
            current_date = datetime.now().strftime("%Y-%m-%d")
            with sqlite3.connect('attendance_system.db') as conn:
                cursor = conn.cursor()
                cursor.execute(
                    "SELECT name, encoding, roll_id FROM students WHERE class_name = ? AND is_authorized = 1",
                    (class_name,))
                rows = cursor.fetchall()
                cursor.execute(
                    "SELECT COUNT(*) FROM attendance WHERE date=? AND time_slot=? AND status='Present' AND roll_id IN (SELECT roll_id FROM students WHERE class_name=?)",
                    (current_date, time_slot, class_name))
                live_count = cursor.fetchone()[0]

            if not rows:
                success, frame = camera.read()
                if not success: break
                frame = cv2.flip(frame, 1)
                cv2.putText(frame, "SYSTEM LOCKED: No Authorized Students found.", (20, 40), cv2.FONT_HERSHEY_SIMPLEX,
                            0.5, (0, 0, 255), 1)
                yield (b'--frame\r\n' b'Content-Type: image/jpeg\r\n\r\n' + cv2.imencode('.jpg', frame)[
                    1].tobytes() + b'\r\n')
                continue

            known_encs = [np.fromstring(r[1], sep=",") for r in rows]
            known_names = [r[0] for r in rows]
            known_rolls = [r[2] for r in rows]
            success, frame = camera.read()
            if not success: break

            frame = cv2.flip(frame, 1)
            cv2.rectangle(frame, (420, 10), (630, 40), (30, 30, 30), cv2.FILLED)
            cv2.putText(frame, f"Verified Today: {live_count}", (430, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (34, 197, 94),
                        1)

            small = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
            rgb_small = cv2.cvtColor(small, cv2.COLOR_BGR2RGB)
            face_locations = face_recognition.face_locations(rgb_small)
            face_encodings = face_recognition.face_encodings(rgb_small, face_locations)

            for (top, right, bottom, left), face_encoding in zip(face_locations, face_encodings):
                name_label, status_text, color = "Unknown Student", "Scanning...", (0, 0, 255)
                if len(known_encs) > 0:
                    matches = face_recognition.compare_faces(known_encs, face_encoding, 0.45)
                    if True in matches:
                        idx = matches.index(True)
                        roll, name = known_rolls[idx], known_names[idx]
                        color = (34, 197, 94)
                        name_label = f"{name} ({roll})"
                        try:
                            with sqlite3.connect('attendance_system.db', timeout=5) as conn:
                                conn.cursor().execute(
                                    "INSERT INTO attendance (roll_id, subject, date, time, time_slot, status) VALUES (?, ?, ?, ?, ?, 'Present')",
                                    (roll, subject, current_date, datetime.now().strftime("%H:%M:%S"), time_slot))
                                conn.commit()
                            status_text = "ATTENDANCE MARKED"
                        except sqlite3.IntegrityError:
                            status_text = "ALREADY MARKED"

                t_orig, r_orig, b_orig, l_orig = top * 4, right * 4, bottom * 4, left * 4
                cv2.rectangle(frame, (l_orig, t_orig), (r_orig, b_orig), color, 2)
                cv2.rectangle(frame, (l_orig, t_orig - 30), (r_orig, t_orig), color, cv2.FILLED)
                cv2.putText(frame, name_label, (l_orig + 6, t_orig - 8), cv2.FONT_HERSHEY_DUPLEX, 0.5, (255, 255, 255),
                            1)
                cv2.rectangle(frame, (l_orig, b_orig), (r_orig, b_orig + 30), color, cv2.FILLED)
                cv2.putText(frame, status_text, (l_orig + 6, b_orig + 20), cv2.FONT_HERSHEY_DUPLEX, 0.5,
                            (255, 255, 255), 1)

            yield (b'--frame\r\n' b'Content-Type: image/jpeg\r\n\r\n' + cv2.imencode('.jpg', frame)[
                1].tobytes() + b'\r\n')
    finally:
        camera.release()


if __name__ == '__main__':
    app.run(debug=True, port=5000)
